import streamlit as st
import torch
import torch.nn as nn
import pandas as pd
import numpy as np
import json
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 参数配置
DATA_COLUMNS = ['evaporation_from_bare_soil_sum',
                'total_precipitation_sum',
                'temperature_2m_max',
                'wind_speed_10m']

INPUT_SIZE = len(DATA_COLUMNS)

# 加载模型参数
with open("models/best_params.json", "r") as f:
    best_params = json.load(f)

# 模型定义
class LSTMRunoffModel(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, dropout=0.1):
        super().__init__()
        self.lstm1 = nn.LSTM(input_size=input_size, hidden_size=hidden_size1, batch_first=True)
        self.lstm2 = nn.LSTM(input_size=hidden_size1, hidden_size=hidden_size2, batch_first=True)
        self.fc = nn.Linear(hidden_size2, 1)
        self.dropout = nn.Dropout(dropout)

    def forward(self, x):
        out, _ = self.lstm1(x)
        out = self.dropout(out)
        out, _ = self.lstm2(out)
        out = self.fc(out)
        return out.squeeze(-1)

@st.cache_resource
def load_model():
    model = LSTMRunoffModel(INPUT_SIZE, best_params['hidden_size1'], best_params['hidden_size2'], best_params['dropout'])
    model.load_state_dict(torch.load("models/best_lstm_model.pth", map_location="cpu"))
    model.eval()
    return model

def create_excel_template(history_days):
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "数据输入"
    ws_guide = wb.create_sheet(title="填写指南")

    headers = ['date'] + DATA_COLUMNS
    ws_data.append(headers)

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = 22 if header != 'date' else 15

    today = datetime.now()
    for i in range(1, history_days + 1):
        date_cell = ws_data[f'A{i+1}']
        date_cell.value = f"{(today + timedelta(days=i-1)).strftime('%Y-%m-%d')}"
        date_cell.number_format = 'yyyy-mm-dd'
        for col_idx in range(2, 6):
            cell = ws_data[f'{get_column_letter(col_idx)}{i+1}']
            dv = DataValidation(type="decimal", operator="greaterThan", formula1="-1000")
            dv.error = '请输入有效的数值！'
            dv.errorTitle = '输入错误'
            ws_data.add_data_validation(dv)
            dv.add(cell)

    ws_data[f'A{history_days + 3}'] = f"⚠️ 注意：请填写完整{history_days}天的连续数据，不可留空"
    ws_data[f'A{history_days + 3}'].font = openpyxl.styles.Font(color="FF0000", bold=True)

    ws_guide['A1'] = "数据填写指南"
    ws_guide['A1'].font = openpyxl.styles.Font(size=16, bold=True)

    field_descriptions = {
        'date': '日期 (格式: YYYY-MM-DD，如2025-06-01)',
        'evaporation_from_bare_soil_sum': '裸土蒸发总量 (单位: mm)',
        'total_precipitation_sum': '总降水量 (单位: mm)',
        'temperature_2m_max': '2米高度最高温度 (单位: °C)',
        'wind_speed_10m': '10米高度风速 (单位: m/s)'
    }

    row = 3
    for field, desc in field_descriptions.items():
        ws_guide[f'A{row}'] = field
        ws_guide[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        ws_guide[f'B{row}'] = desc
        row += 1

    ws_guide['A8'] = "填写要求："
    ws_guide['A8'].font = openpyxl.styles.Font(bold=True)
    ws_guide['B8'] = f"1. 必须提供连续{history_days}天的完整数据"
    ws_guide['B9'] = "2. 日期需按升序排列"
    ws_guide['B10'] = "3. 数值列不可留空，必须为数字"

    ws_guide.column_dimensions['A'].width = 30
    ws_guide.column_dimensions['B'].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def run_forecast_module():
    st.set_page_config(page_title="洪水预报", layout="centered")
    st.title("🌧️ 洪水预报模块")

    st.subheader("参数设置")
    history_days = st.slider("输入历史天数（HISTORY_DAYS）", min_value=7, max_value=30, value=15, step=1)
    forecast_days = st.slider("预测未来天数（FORECAST_DAYS）", min_value=1, max_value=14, value=7, step=1)

    excel_buffer = create_excel_template(history_days)

    st.download_button("📊 下载Excel模板", data=excel_buffer,
                       file_name="data_template.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    csv_template = "date," + ",".join(DATA_COLUMNS) + "\n" + "\n".join(["YYYY-MM-DD,,,," for _ in range(history_days)])
    st.download_button("📄 下载CSV模板", data=csv_template,
                       file_name="data_template.csv", mime="text/csv")

    st.info(f"""💡 请上传或输入连续 **{history_days} 天** 的数据，预测未来 **{forecast_days} 天** 的径流。""")

    manual_input = st.checkbox("手动输入数据")
    df = None

    if manual_input:
        text = st.text_area("输入格式：date,evaporation_from_bare_soil_sum,total_precipitation_sum,temperature_2m_max,wind_speed_10m")
        if text:
            try:
                df = pd.read_csv(StringIO(text)) if ',' in text else pd.read_csv(StringIO(text), sep="\t")
                st.success("✅ 数据读取成功")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"❌ 数据读取失败：{e}")
                return
        else:
            st.warning("请输入数据")
            return
    else:
        uploaded = st.file_uploader("上传 CSV 或 Excel 文件", type=["csv", "xlsx"])
        if uploaded:
            try:
                df = pd.read_csv(uploaded) if uploaded.name.endswith("csv") else pd.read_excel(uploaded)
                st.success("✅ 文件读取成功")
                st.dataframe(df.head())
            except Exception as e:
                st.error(f"❌ 文件读取失败：{e}")
                return
        else:
            st.warning("请上传数据文件")
            return

    if not set(['date'] + DATA_COLUMNS).issubset(df.columns):
        st.error(f"❌ 数据缺失必要列，请确保包含：date + {DATA_COLUMNS}")
        return

    df = df.dropna()
    df['date'] = pd.to_datetime(df['date'])
    features = df[DATA_COLUMNS].values
    dates = df['date'].values

    if len(features) < history_days:
        st.error(f"❌ 数据长度不足 {history_days} 天")
        return

    model = load_model()
    last_history = features[-history_days:]
    last_date = pd.to_datetime(dates[-1])
    predictions, pred_dates = [], []

    for i in range(forecast_days):
        input_tensor = torch.tensor(np.expand_dims(last_history, axis=0), dtype=torch.float32)
        with torch.no_grad():
            output = model(input_tensor)
            prediction = output.numpy()[0, -1]
            predictions.append(prediction)
        new_input = last_history[-1]
        last_history = np.vstack([last_history[1:], new_input])
        pred_dates.append(last_date + timedelta(days=i+1))

    result_df = pd.DataFrame({
        'date': pred_dates,
        'predicted_runoff': predictions
    })

    st.success("✅ 预测完成")
    st.subheader("预测结果")
    st.line_chart(result_df.set_index('date'))
    st.dataframe(result_df)

    st.download_button("📥 下载预测结果", data=result_df.to_csv(index=False).encode('utf-8'), file_name="direct_forecast.csv")

# 请在 app.py 中导入此函数并运行 run_forecast_module()
